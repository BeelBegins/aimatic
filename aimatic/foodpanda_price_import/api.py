import csv
import io

import frappe
import openpyxl
from frappe import _
from frappe.utils import flt, now_datetime
from frappe.utils.file_manager import save_file

from aimatic.shelf_pricing.utils import get_or_create_branch_foodpanda_price_list

# Same role gate shelf_pricing.api uses for any write to a selling price -
# applying an uploaded Foodpanda catalog file is just another price-write
# surface, not a separate permission model.
_ALLOWED_IMPORT_ROLES = {"Buying Price Control", "System Manager"}

_REQUIRED_COLUMNS = ("price", "sku", "name")


def _require_import_permission():
	if not _ALLOWED_IMPORT_ROLES.intersection(frappe.get_roles()):
		frappe.throw(
			_("You need the Buying Price Control role to import a Foodpanda price list."),
			frappe.PermissionError,
		)


def _load_barcode_map():
	barcode_map = {}
	for row in frappe.get_all("Item Barcode", fields=["barcode", "parent"]):
		if row.barcode:
			barcode_map.setdefault(row.barcode.strip(), row.parent)
	return barcode_map


def _resolve_item_code(barcode_map, raw_value):
	"""Foodpanda's own product export prepends one extra leading 0 to every
	barcode it re-exports, on top of whatever is actually on the Item
	Barcode record - confirmed against real ERPNext barcodes that only
	matched once that single leading 0 was stripped. Try the value exactly
	as given first, then with one leading 0 removed."""
	if not raw_value:
		return None

	value = str(raw_value).strip()
	if value in barcode_map:
		return barcode_map[value]

	if value.startswith("0") and value[1:] in barcode_map:
		return barcode_map[value[1:]]

	return None


@frappe.whitelist()
def import_price_list(branch, file_url):
	"""Import a Foodpanda "Products" export (sku, name, price, active,
	barcode 1..barcode N, ...) into `branch`'s own Foodpanda Price List.
	Only `price` (mapped rate=mrp, matching apply_foodpanda_price_update's
	flat FP-price-as-rate convention) is written; every other column is
	informational only. Disabled Items are skipped; barcodes that resolve
	to no Item are collected into a downloadable CSV report rather than
	silently dropped.
	"""
	_require_import_permission()

	if not branch:
		frappe.throw(_("Branch is required."))
	if not file_url:
		frappe.throw(_("Please attach a Foodpanda product export file."))

	price_list = get_or_create_branch_foodpanda_price_list(branch)

	file_doc = frappe.get_doc("File", {"file_url": file_url})
	workbook = openpyxl.load_workbook(io.BytesIO(file_doc.get_content()), data_only=True)
	sheet = workbook.worksheets[0]

	header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
	missing = [col for col in _REQUIRED_COLUMNS if col not in header]
	if missing:
		frappe.throw(_("File is missing required column(s): {0}").format(", ".join(missing)))

	barcode_cols = [i for i, h in enumerate(header) if h and str(h).startswith("barcode")]
	price_col = header.index("price")
	sku_col = header.index("sku")
	name_col = header.index("name")

	barcode_map = _load_barcode_map()
	item_disabled = {i.name: i.disabled for i in frappe.get_all("Item", fields=["name", "disabled"])}
	existing_prices = {
		p.item_code: p
		for p in frappe.get_all(
			"Item Price",
			filters={"price_list": price_list, "selling": 1},
			fields=["name", "item_code", "price_list_rate", "custom_mrp"],
		)
	}

	currency = frappe.db.get_single_value("Global Defaults", "default_currency")
	created = updated = unchanged = skipped_disabled = skipped_bad_price = unmatched = 0
	unmatched_rows = []

	for row in sheet.iter_rows(min_row=2, values_only=True):
		if row is None or row[sku_col] is None:
			continue

		item_code = None
		for col in barcode_cols:
			item_code = _resolve_item_code(barcode_map, row[col])
			if item_code:
				break

		if not item_code:
			unmatched += 1
			unmatched_rows.append([row[sku_col], row[name_col]])
			continue

		if item_disabled.get(item_code):
			skipped_disabled += 1
			continue

		price = flt(row[price_col]) if row[price_col] not in (None, "") else None
		if price is None:
			skipped_bad_price += 1
			continue

		existing = existing_prices.get(item_code)
		if not existing:
			doc = frappe.get_doc(
				{
					"doctype": "Item Price",
					"item_code": item_code,
					"price_list": price_list,
					"selling": 1,
					"currency": currency,
					"price_list_rate": price,
					"custom_mrp": price,
				}
			)
			doc.insert(ignore_permissions=True)
			created += 1
			existing_prices[item_code] = frappe._dict(
				name=doc.name, item_code=item_code, price_list_rate=price, custom_mrp=price
			)
		elif flt(existing.price_list_rate) != price or flt(existing.custom_mrp) != price:
			frappe.db.set_value("Item Price", existing.name, {"price_list_rate": price, "custom_mrp": price})
			updated += 1
			existing.price_list_rate = price
			existing.custom_mrp = price
		else:
			unchanged += 1

	log = frappe.get_doc(
		{
			"doctype": "Foodpanda Price Import Log",
			"branch": branch,
			"price_list": price_list,
			"source_file": file_url,
			"run_by": frappe.session.user,
			"run_datetime": now_datetime(),
			"created_count": created,
			"updated_count": updated,
			"unchanged_count": unchanged,
			"skipped_disabled_count": skipped_disabled,
			"unmatched_count": unmatched,
		}
	)
	log.insert(ignore_permissions=True)

	if unmatched_rows:
		buffer = io.StringIO()
		writer = csv.writer(buffer)
		writer.writerow(["foodpanda_sku", "foodpanda_name"])
		writer.writerows(unmatched_rows)
		saved = save_file(
			f"{log.name}-unmatched.csv",
			buffer.getvalue().encode("utf-8"),
			"Foodpanda Price Import Log",
			log.name,
			is_private=1,
		)
		log.db_set("unmatched_report", saved.file_url)

	return {
		"log": log.name,
		"price_list": price_list,
		"created": created,
		"updated": updated,
		"unchanged": unchanged,
		"skipped_disabled": skipped_disabled,
		"unmatched": unmatched,
		"unmatched_report": log.unmatched_report,
	}
