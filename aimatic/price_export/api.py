import io

import frappe
import openpyxl
from frappe.utils import flt, now_datetime, parse_json

from aimatic.branch_management.utils import get_branch_defaults
from aimatic.shelf_pricing.utils import get_or_create_branch_foodpanda_price_list

# Cost price (excl/incl tax) exposes actual purchase cost, unlike the plain
# selling-price read in price_check/shelf_pricing - gate it the same way
# shelf_pricing.api/foodpanda_price_import.api gate any cost-sensitive
# read/write, not the broader Item-read check price_check uses. This is the
# entry-point security boundary for the whole module (matching vendor_
# performance/sales_dashboard's own "gate at the API, query freely
# underneath" convention) - every helper below reads with ignore_permissions
# since Buying Price Control has no standing DocPerm on Item Price/Warehouse/
# Bin/Item Barcode of its own, only the Branch/Item read+report grants added
# for this report (see patches.repair_item_custom_docperms and hooks.py's
# Custom DocPerm fixture block).
ALLOWED_EXPORT_ROLES = {"Buying Price Control", "System Manager"}

_CHUNK_SIZE = 2000
_MAX_GRID_PRICE_UPDATES = 20000
_MAX_BRANCH_SHEET_BYTES = 25 * 1024 * 1024
_BRANCH_SHEET_PRICE_HEADERS = (
    "foodpanda price (editable)",
    "foodpanda price",
)


def require_export_permission():
    if not ALLOWED_EXPORT_ROLES.intersection(frappe.get_roles()):
        frappe.throw(
            frappe._("You need the Buying Price Control role to view the branch price sheet."),
            frappe.PermissionError,
        )


def _chunks(seq, size):
    for i in range(0, len(seq), size):
        yield seq[i : i + size]


def _resolve_branch_price_lists(branch):
    """Read-only - deliberately does not create either list (unlike the
    Purchase Receipt shelf-pricing flow), matching get_current_branch_sale_price/
    get_current_foodpanda_price's own "a plain read must not create" rule."""
    selling_price_list = frappe.db.get_value(
        "Branch", branch, "default_selling_price_list"
    ) or frappe.db.get_single_value("Selling Settings", "selling_price_list")
    foodpanda_price_list = frappe.db.get_value("Branch", branch, "default_foodpanda_price_list")
    return selling_price_list, foodpanda_price_list


def _fetch_price_map_by_uom(price_list, item_codes):
    """{item_code: {uom: {rate, mrp}}} - an item can carry more than one
    selling Item Price row on the same price list at different UOMs (e.g. a
    Pcs rate and a separate, independently-entered Box rate). Collapsing
    these into one number per item (the previous behaviour) picked whichever
    row the query happened to return last, which is how a real Box rate
    (3050) once got treated as the Pcs selling price and then multiplied by
    the Box conversion factor again (3050 x 5) - see get_branch_price_sheet_rows.
    """
    if not price_list:
        return {}
    prices = {}
    for chunk in _chunks(item_codes, _CHUNK_SIZE):
        for row in frappe.get_all(
            "Item Price",
            filters={"price_list": price_list, "selling": 1, "item_code": ("in", chunk)},
            fields=["item_code", "uom", "price_list_rate", "custom_mrp"],
            ignore_permissions=True,
        ):
            prices.setdefault(row.item_code, {})[row.uom] = {
                "rate": flt(row.price_list_rate),
                "mrp": flt(row.custom_mrp),
            }
    return prices


def _fetch_barcode_map(item_codes):
    """{item_code: [barcode, ...]} - kept as a list so the report can render
    one column per barcode (barcode1, barcode2, ...) instead of a single
    comma-joined string."""
    barcodes = {}
    for chunk in _chunks(item_codes, _CHUNK_SIZE):
        for row in frappe.get_all(
            "Item Barcode",
            filters={"parent": ("in", chunk)},
            fields=["parent", "barcode", "idx"],
            order_by="idx asc",
            ignore_permissions=True,
        ):
            if row.barcode:
                barcodes.setdefault(row.parent, []).append(row.barcode)
    return barcodes


def _fetch_box_conversion_map(item_codes):
    """conversion_factor of each item's own 'Box' UOM Conversion Detail row,
    if one exists - most items have none (see the shelf-pricing/box-price
    clarification: only a real per-item Box UOM entry counts, no synthetic
    default)."""
    factors = {}
    for chunk in _chunks(item_codes, _CHUNK_SIZE):
        for row in frappe.get_all(
            "UOM Conversion Detail",
            filters={"parent": ("in", chunk), "parenttype": "Item", "uom": "Box"},
            fields=["parent", "conversion_factor"],
            ignore_permissions=True,
        ):
            factors[row.parent] = flt(row.conversion_factor)
    return factors


def _fetch_stock_map(branch, item_codes):
    warehouses = frappe.get_all(
        "Warehouse",
        filters={"custom_branch": branch, "disabled": 0},
        pluck="name",
        ignore_permissions=True,
    )
    if not warehouses:
        return {}
    stock = {}
    for chunk in _chunks(item_codes, _CHUNK_SIZE):
        rows = frappe.db.sql(
            """
            SELECT item_code, SUM(actual_qty) AS qty
            FROM `tabBin`
            WHERE warehouse IN %(warehouses)s AND item_code IN %(items)s
            GROUP BY item_code
            """,
            {"warehouses": warehouses, "items": chunk},
            as_dict=True,
        )
        for row in rows:
            stock[row.item_code] = flt(row.qty)
    return stock


def _fetch_available_stock_map(branch, item_codes):
    """Sellable Foodpanda quantity in the branch's finished-goods warehouse.

    This deliberately differs from the report's Stock In Hand total, which
    sums every enabled warehouse tagged to the branch. Foodpanda availability
    must exclude reserved stock and must never include a rejected warehouse.
    """
    warehouse = get_branch_defaults(branch).get("finished_goods_warehouse")
    if not warehouse:
        return {}

    available = {}
    for chunk in _chunks(item_codes, _CHUNK_SIZE):
        for row in frappe.get_all(
            "Bin",
            filters={"warehouse": warehouse, "item_code": ("in", chunk)},
            fields=["item_code", "actual_qty", "reserved_qty"],
            ignore_permissions=True,
        ):
            available[row.item_code] = max(
                flt(row.actual_qty) - flt(row.reserved_qty),
                0,
            )
    return available


_SOURCE_DOCTYPES = ("Purchase Receipt", "Purchase Invoice")


def _fetch_purchase_cost_rows(item_codes):
    """Cost candidates from the custom_price_without_taxes/custom_price_after_taxes
    fields on submitted, non-return Purchase Receipt/Purchase Invoice rows.
    These custom fields are only populated for documents entered through this
    app's own purchase forms, so coverage is partial - most items never get a
    row from this source and must fall back to _fetch_stock_entry_cost_rows.
    """
    rows_out = []
    for chunk in _chunks(item_codes, _CHUNK_SIZE):
        for source_doctype in _SOURCE_DOCTYPES:
            child_table = f"`tab{source_doctype} Item`"
            parent_table = f"`tab{source_doctype}`"
            rows = frappe.db.sql(
                f"""
                SELECT
                    child.item_code AS item_code,
                    child.custom_price_without_taxes AS excl_tax,
                    child.custom_price_after_taxes AS incl_tax,
                    parent.posting_date AS posting_date,
                    parent.posting_time AS posting_time,
                    child.creation AS creation
                FROM {child_table} child
                INNER JOIN {parent_table} parent ON parent.name = child.parent
                WHERE parent.docstatus = 1
                  AND parent.is_return = 0
                  AND child.item_code IN %(items)s
                  AND (child.custom_price_without_taxes > 0 OR child.custom_price_after_taxes > 0)
                """,
                {"items": chunk},
                as_dict=True,
            )
            rows_out.extend(rows)
    return rows_out


def _fetch_stock_entry_cost_rows(item_codes):
    """Cost candidates from plain Stock Entry rows (Material Receipt/Material
    Issue/Transfer, whichever actually carries a rate). Stock Entry has no
    per-line GST breakdown, and its valuation_rate is not tax-inclusive. This
    matters especially for the iPOS opening migration: CurCost was explicitly
    reverse-calculated to tax-exclusive before both basic_rate and
    valuation_rate were stored, while the removed GST was posted separately
    through a Journal Entry. Reconstruct the display-only inclusive rate from
    that exclusive rate and the same Item FBR category tax rate used by the
    migration; never mistake valuation_rate for GST-inclusive cost.
    """
    rows_out = []
    for chunk in _chunks(item_codes, _CHUNK_SIZE):
        rows = frappe.db.sql(
            """
            SELECT
                child.item_code AS item_code,
                child.basic_rate AS basic_rate,
                child.valuation_rate AS valuation_rate,
                COALESCE(tax_category.tax_rate, 0) AS tax_rate,
                parent.posting_date AS posting_date,
                parent.posting_time AS posting_time,
                child.creation AS creation
            FROM `tabStock Entry Detail` child
            INNER JOIN `tabStock Entry` parent ON parent.name = child.parent
            INNER JOIN `tabItem` item ON item.name = child.item_code
            LEFT JOIN `tabFBR Tax Category` tax_category
              ON tax_category.name = item.custom_fbr_tax_category
            WHERE parent.docstatus = 1
              AND child.item_code IN %(items)s
              AND (child.basic_rate > 0 OR child.valuation_rate > 0)
            """,
            {"items": chunk},
            as_dict=True,
        )
        for row in rows:
            exclusive_rate = flt(row.basic_rate) or flt(row.valuation_rate)
            rows_out.append(
                frappe._dict(
                    item_code=row.item_code,
                    excl_tax=exclusive_rate,
                    incl_tax=_inclusive_rate_from_exclusive(exclusive_rate, row.tax_rate),
                    posting_date=row.posting_date,
                    posting_time=row.posting_time,
                    creation=row.creation,
                )
            )
    return rows_out


def _inclusive_rate_from_exclusive(exclusive_rate, tax_rate):
    """Inverse of ipos_data_migration.exclusive_rate for report display."""
    exclusive_rate = flt(exclusive_rate)
    tax_rate = flt(tax_rate)
    if exclusive_rate <= 0 or tax_rate <= 0:
        return exclusive_rate
    return round(exclusive_rate * (100 + tax_rate) / 100, 4)


def _fetch_cost_price_map(item_codes):
    """Latest priced row per item across Purchase Receipt/Purchase Invoice
    (custom tax-inclusive fields) and Stock Entry (basic_rate/valuation_rate)
    combined - site-wide, not branch-filtered, matching item_pricing's own
    Item.custom_latest_price_incl_taxes convention (a single global "current
    cost", not a per-branch one). Excl-tax and incl-tax values are read off
    the same winning row so the pair is always internally consistent; the two
    sources are compared on equal footing by posting_date/posting_time/creation,
    so whichever document was actually posted most recently wins regardless
    of which source it came from.
    """
    best: dict[str, dict] = {}

    all_rows = _fetch_purchase_cost_rows(item_codes) + _fetch_stock_entry_cost_rows(item_codes)
    for row in all_rows:
        current = best.get(row.item_code)
        key = (row.posting_date, row.posting_time, row.creation)
        if current is None or key > current["_key"]:
            best[row.item_code] = {
                "excl_tax": flt(row.excl_tax),
                "incl_tax": flt(row.incl_tax),
                "_key": key,
            }

    return {code: {"excl_tax": v["excl_tax"], "incl_tax": v["incl_tax"]} for code, v in best.items()}


def get_branch_price_sheet_rows(branch):
    """Core computation backing the 'Branch Price Sheet' Script Report
    (aimatic/aimatic/report/branch_price_sheet/): every active sales Item's
    branch-scoped pricing/stock snapshot - current selling price and MRP
    (from this branch's Selling Price List), Foodpanda price (from this
    branch's Foodpanda Price List), current stock in hand (this branch's own
    warehouses), barcodes, UOM, latest purchase
    cost price excl/incl taxes (site-wide, see _fetch_cost_price_map), and
    Box price (current selling price x the item's own Box UOM conversion
    factor, where one exists). Returns a list of dicts keyed by column
    fieldname, in Item Code order.
    """
    items = frappe.get_all(
        "Item",
        filters={"disabled": 0, "is_sales_item": 1},
        fields=["item_code", "item_name", "stock_uom"],
        order_by="item_code asc",
        ignore_permissions=True,
    )
    item_codes = [i.item_code for i in items]

    selling_price_list, foodpanda_price_list = _resolve_branch_price_lists(branch)
    selling_prices_by_uom = _fetch_price_map_by_uom(selling_price_list, item_codes)
    foodpanda_prices_by_uom = _fetch_price_map_by_uom(foodpanda_price_list, item_codes)
    barcodes = _fetch_barcode_map(item_codes)
    box_factors = _fetch_box_conversion_map(item_codes)
    stock = _fetch_stock_map(branch, item_codes)
    available_stock = _fetch_available_stock_map(branch, item_codes)
    cost_prices = _fetch_cost_price_map(item_codes)

    rows = []
    for item in items:
        code = item.item_code
        cost = cost_prices.get(code, {})
        box_factor = box_factors.get(code)

        item_prices_by_uom = selling_prices_by_uom.get(code, {})
        # An item can have its own selling rate per UOM (a Pcs row and a
        # separately-entered Box row). Prefer the row matching the item's
        # stock UOM for "current selling price"; only fall back to whatever
        # single row exists when there's no exact stock-UOM match.
        if item.stock_uom in item_prices_by_uom:
            selling_price_row = item_prices_by_uom[item.stock_uom]
        elif len(item_prices_by_uom) == 1:
            selling_price_row = next(iter(item_prices_by_uom.values()))
        else:
            selling_price_row = None
        selling_price = selling_price_row["rate"] if selling_price_row else None
        mrp = selling_price_row["mrp"] if selling_price_row else None

        foodpanda_prices_for_item = foodpanda_prices_by_uom.get(code, {})
        if item.stock_uom in foodpanda_prices_for_item:
            foodpanda_price = foodpanda_prices_for_item[item.stock_uom]["rate"]
        elif len(foodpanda_prices_for_item) == 1:
            foodpanda_price = next(iter(foodpanda_prices_for_item.values()))["rate"]
        else:
            foodpanda_price = None

        # If the item already carries its own explicit "Box" Item Price row,
        # that IS the box price - use it directly. Multiplying it by the Box
        # conversion factor again double-counts it (a real bug: an item
        # priced at 3050/Box was being shown as 3050 x 5 here). Only derive
        # Box price by multiplication when no explicit Box row exists.
        if "Box" in item_prices_by_uom:
            box_price = item_prices_by_uom["Box"]["rate"]
        elif box_factor and selling_price:
            box_price = selling_price * box_factor
        else:
            box_price = None

        row = {
            "item_code": code,
            "item_name": item.item_name,
            "uom": item.stock_uom,
            "selling_price": selling_price,
            "mrp": mrp,
            "foodpanda_price": foodpanda_price,
            "foodpanda_active": 1 if available_stock.get(code, 0) > 0 else 0,
            "available_qty": available_stock.get(code, 0),
            "stock_in_hand": stock.get(code, 0),
            "cost_price_excl_tax": cost.get("excl_tax"),
            "cost_price_incl_tax": cost.get("incl_tax"),
            "box_price": box_price,
            "_barcodes": barcodes.get(code, []),
        }
        rows.append(row)

    return rows


def _normalize_foodpanda_price_updates(updates):
    if isinstance(updates, str):
        updates = parse_json(updates)
    if not isinstance(updates, list):
        raise ValueError("Foodpanda price updates must be a list.")
    if len(updates) > _MAX_GRID_PRICE_UPDATES:
        raise ValueError(
            f"A maximum of {_MAX_GRID_PRICE_UPDATES} Foodpanda prices can be saved at once."
        )

    normalized = {}
    for row in updates:
        if not isinstance(row, dict):
            raise ValueError("Every Foodpanda price update must be an object.")
        item_code = str(row.get("item_code") or "").strip()
        price = flt(row.get("price"))
        old_price = flt(row.get("old_price"))
        if not item_code:
            raise ValueError("Every Foodpanda price update requires an Item Code.")
        if price <= 0:
            raise ValueError(f"Foodpanda price for {item_code} must be greater than zero.")
        normalized[item_code] = {"price": price, "old_price": old_price}
    return normalized


def _select_foodpanda_price_row(rows, stock_uom):
    exact = next((row for row in rows if row.uom == stock_uom), None)
    if exact:
        return exact
    if len(rows) == 1:
        return rows[0]
    return None


def _apply_foodpanda_price_updates(branch, normalized, source_file=None, skip_invalid=False):
    """Apply validated item/price pairs to one branch Foodpanda Price List."""
    normalized = dict(normalized)
    item_codes = list(normalized)
    item_rows = frappe.get_all(
        "Item",
        filters={"name": ("in", item_codes)},
        fields=["name", "stock_uom", "disabled", "is_sales_item"],
        ignore_permissions=True,
    )
    items = {row.name: row for row in item_rows}
    invalid = [
        code
        for code in item_codes
        if code not in items or items[code].disabled or not items[code].is_sales_item
    ]
    if invalid:
        if not skip_invalid:
            frappe.throw(
                frappe._("These Items are missing, disabled, or not sales items: {0}").format(
                    ", ".join(invalid[:20])
                )
            )
        for code in invalid:
            normalized.pop(code, None)
        item_codes = list(normalized)
        if not item_codes:
            return {
                "log": None,
                "price_list": None,
                "created": 0,
                "updated": 0,
                "unchanged": 0,
                "skipped_invalid": len(invalid),
            }

    price_list = get_or_create_branch_foodpanda_price_list(branch)
    existing_by_item = {}
    for row in frappe.get_all(
        "Item Price",
        filters={"price_list": price_list, "selling": 1, "item_code": ("in", item_codes)},
        fields=["name", "item_code", "uom", "price_list_rate", "custom_mrp"],
        order_by="modified desc",
        ignore_permissions=True,
    ):
        existing_by_item.setdefault(row.item_code, []).append(row)

    selected_rows = {
        code: _select_foodpanda_price_row(existing_by_item.get(code, []), items[code].stock_uom)
        for code in item_codes
    }
    conflicts = []
    for code, values in normalized.items():
        current_row = selected_rows[code]
        current_price = flt(current_row.price_list_rate) if current_row else 0
        if "old_price" in values and current_price != values["old_price"]:
            conflicts.append(code)
    if conflicts:
        frappe.throw(
            frappe._(
                "Foodpanda prices changed after this report loaded. Refresh and retry these Items: {0}"
            ).format(", ".join(conflicts[:20]))
        )

    created = updated = unchanged = 0
    currency = frappe.db.get_value("Price List", price_list, "currency") or frappe.db.get_single_value(
        "Global Defaults", "default_currency"
    )
    for code, values in normalized.items():
        price = values["price"]
        existing = selected_rows[code]
        if existing:
            if flt(existing.price_list_rate) == price and flt(existing.custom_mrp) == price:
                unchanged += 1
                continue
            frappe.db.set_value(
                "Item Price",
                existing.name,
                {"price_list_rate": price, "custom_mrp": price},
            )
            updated += 1
            continue

        frappe.get_doc(
            {
                "doctype": "Item Price",
                "item_code": code,
                "uom": items[code].stock_uom,
                "price_list": price_list,
                "selling": 1,
                "currency": currency,
                "price_list_rate": price,
                "custom_mrp": price,
            }
        ).insert(ignore_permissions=True)
        created += 1

    log_values = {
        "doctype": "Foodpanda Price Import Log",
        "branch": branch,
        "price_list": price_list,
        "run_by": frappe.session.user,
        "run_datetime": now_datetime(),
        "created_count": created,
        "updated_count": updated,
        "unchanged_count": unchanged,
        "skipped_disabled_count": len(invalid),
    }
    if source_file:
        log_values["source_file"] = source_file
    log = frappe.get_doc(log_values)
    log.insert(ignore_permissions=True)
    return {
        "log": log.name,
        "price_list": price_list,
        "created": created,
        "updated": updated,
        "unchanged": unchanged,
        "skipped_invalid": len(invalid),
    }


def _normalize_branch_sheet_header(value):
    return " ".join(str(value or "").strip().lower().split())


def _extract_branch_price_sheet_updates(sheet):
    """Read the standard Frappe report Excel shape, with optional filter rows."""
    header_row = item_column = price_column = None
    for row_number, values in enumerate(
        sheet.iter_rows(min_row=1, max_row=50, values_only=True), start=1
    ):
        headers = [_normalize_branch_sheet_header(value) for value in values]
        if "item code" not in headers:
            continue
        matching_price_headers = [
            header for header in _BRANCH_SHEET_PRICE_HEADERS if header in headers
        ]
        if not matching_price_headers:
            continue
        header_row = row_number
        item_column = headers.index("item code")
        price_column = headers.index(matching_price_headers[0])
        break

    if header_row is None:
        raise ValueError(
            "This is not a Branch Price Sheet Excel file. Required columns: "
            "Item Code and Foodpanda Price (Editable)."
        )

    updates = {}
    duplicates = set()
    skipped_blank_price = skipped_bad_price = 0
    for values in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        item_value = values[item_column] if item_column < len(values) else None
        price_value = values[price_column] if price_column < len(values) else None
        item_code = str(item_value or "").strip()
        if not item_code:
            continue
        if price_value in (None, ""):
            skipped_blank_price += 1
            continue
        price = flt(str(price_value).replace(",", ""))
        if price <= 0:
            skipped_bad_price += 1
            continue
        if item_code in updates:
            duplicates.add(item_code)
        updates[item_code] = {"price": price}

    if duplicates:
        raise ValueError(
            "Duplicate Item Codes are not allowed in the workbook: "
            + ", ".join(sorted(duplicates)[:20])
        )
    if len(updates) > _MAX_GRID_PRICE_UPDATES:
        raise ValueError(
            f"A maximum of {_MAX_GRID_PRICE_UPDATES} Foodpanda prices can be imported at once."
        )
    return updates, {
        "workbook_rows": len(updates) + skipped_blank_price + skipped_bad_price,
        "accepted_rows": len(updates),
        "skipped_blank_price": skipped_blank_price,
        "skipped_bad_price": skipped_bad_price,
    }


def _require_branch_price_update_access(branch):
    require_export_permission()
    if not branch:
        frappe.throw(frappe._("Branch is required."))
    if not frappe.has_permission("Branch", ptype="read", doc=branch):
        frappe.throw(frappe._("Not permitted to update this branch."), frappe.PermissionError)


@frappe.whitelist()
def save_foodpanda_grid_prices(branch, updates):
    """Save changed report-grid prices to one branch's Foodpanda Price List.

    Stock, active, and quantity are never accepted from the browser: those
    remain server-derived read-only facts. Each row carries the price seen
    when the report loaded, so a concurrent newer edit is rejected instead
    of being silently overwritten.
    """
    _require_branch_price_update_access(branch)
    try:
        normalized = _normalize_foodpanda_price_updates(updates)
    except ValueError as error:
        frappe.throw(frappe._(str(error)))
    if not normalized:
        return {"created": 0, "updated": 0, "unchanged": 0}
    return _apply_foodpanda_price_updates(branch, normalized)


@frappe.whitelist()
def import_branch_price_sheet(branch, file_url):
    """Round-trip a standard Branch Price Sheet Excel export.

    Only Item Code and Foodpanda Price are read. Every stock, availability,
    selling-price, MRP, cost, and barcode column is informational and ignored.
    """
    _require_branch_price_update_access(branch)
    if not file_url:
        frappe.throw(frappe._("Please attach the updated Branch Price Sheet Excel file."))

    file_doc = frappe.get_doc("File", {"file_url": file_url})
    file_name = (file_doc.file_name or file_url).lower()
    if not file_name.endswith(".xlsx"):
        frappe.throw(frappe._("Please upload an .xlsx file exported from Branch Price Sheet."))
    content = file_doc.get_content()
    if len(content) > _MAX_BRANCH_SHEET_BYTES:
        frappe.throw(frappe._("The Excel file is too large. Maximum allowed size is 25 MB."))

    try:
        workbook = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        try:
            updates, import_stats = _extract_branch_price_sheet_updates(workbook.worksheets[0])
        finally:
            workbook.close()
    except ValueError as error:
        frappe.throw(frappe._(str(error)))
    except Exception:
        frappe.throw(frappe._("Could not read this Excel file. Export it again and retry."))

    if not updates:
        frappe.throw(frappe._("No positive Foodpanda prices were found in the Excel file."))

    result = _apply_foodpanda_price_updates(
        branch,
        updates,
        source_file=file_url,
        skip_invalid=True,
    )
    result.update(import_stats)
    return result
