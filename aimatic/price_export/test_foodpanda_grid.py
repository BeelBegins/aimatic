import unittest
from types import SimpleNamespace
from unittest.mock import Mock, patch

from openpyxl import Workbook

from aimatic.price_export.api import (
    _extract_branch_price_sheet_updates,
    _normalize_foodpanda_price_updates,
    _select_foodpanda_price_row,
)


class TestFoodpandaPriceGrid(unittest.TestCase):
    def test_extracts_round_trip_excel_with_optional_filter_rows(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Branch", "S1 - Ghouri Town VIP"])
        sheet.append(["Foodpanda Availability", "In Stock"])
        sheet.append([])
        sheet.append(
            [
                "Item Code",
                "Current Selling Price",
                "Foodpanda Price (Editable)",
                "FP Active",
                "FP Available Qty",
            ]
        )
        sheet.append(["ITEM-1", 100, 125.5, 1, 10])
        sheet.append(["ITEM-2", 200, None, 1, 20])
        sheet.append(["ITEM-3", 300, 0, 0, 0])

        updates, stats = _extract_branch_price_sheet_updates(sheet)
        workbook.close()

        self.assertEqual(updates, {"ITEM-1": {"price": 125.5}})
        self.assertEqual(stats["accepted_rows"], 1)
        self.assertEqual(stats["skipped_blank_price"], 1)
        self.assertEqual(stats["skipped_bad_price"], 1)

    def test_round_trip_excel_rejects_duplicate_item_codes(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Item Code", "Foodpanda Price"])
        sheet.append(["ITEM-1", 100])
        sheet.append(["ITEM-1", 110])

        with self.assertRaisesRegex(ValueError, "Duplicate Item Codes"):
            _extract_branch_price_sheet_updates(sheet)
        workbook.close()

    def test_round_trip_excel_requires_report_headers(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["SKU", "Price"])

        with self.assertRaisesRegex(ValueError, "not a Branch Price Sheet"):
            _extract_branch_price_sheet_updates(sheet)
        workbook.close()

    def test_normalizes_updates_and_last_duplicate_wins(self):
        result = _normalize_foodpanda_price_updates(
            [
                {"item_code": " ITEM-1 ", "old_price": 100, "price": 110},
                {"item_code": "ITEM-1", "old_price": 100, "price": 115},
            ]
        )

        self.assertEqual(result, {"ITEM-1": {"old_price": 100.0, "price": 115.0}})

    def test_accepts_json_request_payload(self):
        result = _normalize_foodpanda_price_updates(
            '[{"item_code":"ITEM-1","old_price":0,"price":99.5}]'
        )

        self.assertEqual(result, {"ITEM-1": {"old_price": 0.0, "price": 99.5}})

    def test_rejects_zero_price(self):
        with self.assertRaisesRegex(ValueError, "greater than zero"):
            _normalize_foodpanda_price_updates(
                [{"item_code": "ITEM-1", "old_price": 100, "price": 0}]
            )

    def test_selects_exact_stock_uom_from_multiple_rows(self):
        rows = [
            SimpleNamespace(name="PRICE-BOX", uom="Box"),
            SimpleNamespace(name="PRICE-PCS", uom="Pcs"),
        ]

        selected = _select_foodpanda_price_row(rows, "Pcs")

        self.assertEqual(selected.name, "PRICE-PCS")

    def test_uses_single_row_fallback_but_not_ambiguous_rows(self):
        only_row = SimpleNamespace(name="PRICE-ONLY", uom="Box")
        self.assertIs(_select_foodpanda_price_row([only_row], "Pcs"), only_row)
        self.assertIsNone(
            _select_foodpanda_price_row(
                [
                    SimpleNamespace(name="PRICE-BOX", uom="Box"),
                    SimpleNamespace(name="PRICE-PACK", uom="Pack"),
                ],
                "Pcs",
            )
        )

    @patch("aimatic.price_export.api.get_or_create_branch_foodpanda_price_list")
    @patch("aimatic.price_export.api.require_export_permission")
    @patch("aimatic.price_export.api.frappe")
    def test_save_creates_only_price_fields_in_branch_foodpanda_list(
        self, frappe, _require_permission, get_price_list
    ):
        from aimatic.price_export.api import save_foodpanda_grid_prices

        get_price_list.return_value = "Branch One Foodpanda Price List"
        frappe.has_permission.return_value = True
        frappe.get_all.side_effect = [
            [SimpleNamespace(name="ITEM-1", stock_uom="Pcs", disabled=0, is_sales_item=1)],
            [],
        ]
        frappe.db.get_value.return_value = "PKR"
        frappe.session.user = "price.manager@example.com"
        price_doc = Mock(name="price_doc")
        price_doc.name = "ITEM-PRICE-1"
        log_doc = Mock(name="log_doc")
        log_doc.name = "FP-LOG-1"
        frappe.get_doc.side_effect = [price_doc, log_doc]

        with patch("aimatic.price_export.api.now_datetime", return_value="2026-08-04 12:00:00"):
            result = save_foodpanda_grid_prices(
                "Branch One",
                [
                    {
                        "item_code": "ITEM-1",
                        "old_price": 0,
                        "price": 125,
                        "active": 0,
                        "quantity": 999,
                    }
                ],
            )

        price_values = frappe.get_doc.call_args_list[0].args[0]
        self.assertEqual(price_values["price_list"], "Branch One Foodpanda Price List")
        self.assertEqual(price_values["price_list_rate"], 125)
        self.assertEqual(price_values["custom_mrp"], 125)
        self.assertNotIn("active", price_values)
        self.assertNotIn("quantity", price_values)
        price_doc.insert.assert_called_once_with(ignore_permissions=True)
        self.assertEqual(result["created"], 1)
        self.assertEqual(result["log"], "FP-LOG-1")

    @patch("aimatic.price_export.api.get_or_create_branch_foodpanda_price_list")
    @patch("aimatic.price_export.api.require_export_permission")
    @patch("aimatic.price_export.api.frappe")
    def test_save_rejects_stale_price(self, frappe, _require_permission, get_price_list):
        from aimatic.price_export.api import save_foodpanda_grid_prices

        get_price_list.return_value = "Branch One Foodpanda Price List"
        frappe.has_permission.return_value = True
        frappe._.side_effect = lambda message: message
        frappe.throw.side_effect = RuntimeError
        frappe.get_all.side_effect = [
            [SimpleNamespace(name="ITEM-1", stock_uom="Pcs", disabled=0, is_sales_item=1)],
            [
                SimpleNamespace(
                    name="ITEM-PRICE-1",
                    item_code="ITEM-1",
                    uom="Pcs",
                    price_list_rate=120,
                    custom_mrp=120,
                )
            ],
        ]

        with self.assertRaises(RuntimeError):
            save_foodpanda_grid_prices(
                "Branch One",
                [{"item_code": "ITEM-1", "old_price": 100, "price": 130}],
            )

        frappe.db.set_value.assert_not_called()
        frappe.get_doc.assert_not_called()

    @patch("aimatic.price_export.api.get_or_create_branch_foodpanda_price_list")
    @patch("aimatic.price_export.api.frappe")
    def test_excel_apply_updates_only_foodpanda_price_and_logs_source_file(
        self, frappe, get_price_list
    ):
        from aimatic.price_export.api import _apply_foodpanda_price_updates

        get_price_list.return_value = "Branch One Foodpanda Price List"
        frappe.get_all.side_effect = [
            [SimpleNamespace(name="ITEM-1", stock_uom="Pcs", disabled=0, is_sales_item=1)],
            [
                SimpleNamespace(
                    name="ITEM-PRICE-1",
                    item_code="ITEM-1",
                    uom="Pcs",
                    price_list_rate=100,
                    custom_mrp=100,
                )
            ],
        ]
        frappe.db.get_value.return_value = "PKR"
        frappe.session.user = "price.manager@example.com"
        log_doc = Mock(name="log_doc")
        log_doc.name = "FP-LOG-2"
        frappe.get_doc.return_value = log_doc

        with patch("aimatic.price_export.api.now_datetime", return_value="2026-08-04 12:00:00"):
            result = _apply_foodpanda_price_updates(
                "Branch One",
                {"ITEM-1": {"price": 125}},
                source_file="/private/files/branch-price-sheet.xlsx",
                skip_invalid=True,
            )

        frappe.db.set_value.assert_called_once_with(
            "Item Price",
            "ITEM-PRICE-1",
            {"price_list_rate": 125, "custom_mrp": 125},
        )
        log_values = frappe.get_doc.call_args.args[0]
        self.assertEqual(log_values["source_file"], "/private/files/branch-price-sheet.xlsx")
        self.assertEqual(log_values["updated_count"], 1)
        log_doc.insert.assert_called_once_with(ignore_permissions=True)
        self.assertEqual(result["updated"], 1)


if __name__ == "__main__":
    unittest.main()
