import hashlib
import io
import json
import uuid
from math import floor

import frappe
from frappe import _
from frappe.utils import flt, now_datetime

from aimatic.branch_management.utils import get_branch_defaults
from aimatic.foodpanda_integration import client
from aimatic.foodpanda_integration.client import FoodpandaAPIError
from aimatic.shelf_pricing.utils import get_or_create_branch_foodpanda_price_list

_ADD_PRODUCTS_PATH = "/v2/chains/{chain_id}/catalog"
_UPDATE_PRODUCTS_PATH = "/v2/chains/{chain_id}/vendors/{vendor_id}/catalog"
_CATALOG_PATH = "/v2/chains/{chain_id}/vendors/{vendor_id}/catalog"
_CATEGORIES_PATH = "/v2/chains/{chain_id}/vendors/{vendor_id}/categories"
_EXPORT_PATH = "/v2/chains/{chain_id}/vendors/{vendor_id}/catalog/export"
_JOB_STATUS_PATH = "/v2/chains/{chain_id}/catalog/jobs/{job_id}"

_JOB_CACHE_PREFIX = "aimatic:foodpanda:bulkexport:"
_JOB_CACHE_TTL = 24 * 60 * 60


def _get_item_price(item_code, branch):
	price_list = get_or_create_branch_foodpanda_price_list(branch)
	rate = frappe.db.get_value(
		"Item Price", {"item_code": item_code, "price_list": price_list, "selling": 1}, "price_list_rate"
	)
	return flt(rate, 2)


def _get_stock(item_code, branch):
	"""Returns (quantity, active) for one branch's Bin - quantity is the real
	sellable stock number Foodpanda's `quantity` field expects, active is
	just quantity > 0. A single Bin lookup backs both."""
	warehouse = get_branch_defaults(branch).get("finished_goods_warehouse")
	if not warehouse:
		return 0, False
	bin_row = frappe.db.get_value(
		"Bin", {"item_code": item_code, "warehouse": warehouse}, ["actual_qty", "reserved_qty"], as_dict=True
	) or {}
	quantity = max(flt(bin_row.get("actual_qty")) - flt(bin_row.get("reserved_qty")), 0)
	return quantity, quantity > 0


def _get_barcode(item_code):
	return frappe.db.get_value("Item Barcode", {"parent": item_code}, "barcode")


def _is_foodpanda_barcode(barcode):
	"""Foodpanda PUT rejects non-digit / >14-digit barcode values."""
	value = str(barcode or "").strip()
	return bool(value) and value.isdigit() and len(value) <= 14


def _get_foodpanda_barcode(item_code):
	"""Prefer a digit-only barcode Foodpanda will accept; else None."""
	barcodes = frappe.get_all(
		"Item Barcode", filters={"parent": item_code}, pluck="barcode"
	)
	for barcode in barcodes:
		if _is_foodpanda_barcode(barcode):
			return str(barcode).strip()
	return None



def _barcode_variants(barcode):
	"""Exact barcode plus Foodpanda/ERPNext numeric padding variants.

	Foodpanda often prefixes a leading 0 (e.g. 08851932354998 vs local
	8851932354998). Match both directions: strip zeros, drop one leading
	zero, and common GTIN-13/14 zero-padding.
	"""
	value = str(barcode or "").strip()
	if not value:
		return []
	variants = {value}
	if value.isdigit():
		stripped = value.lstrip("0") or "0"
		variants.add(stripped)
		if value.startswith("0") and len(value) > 1:
			variants.add(value[1:])
		else:
			variants.add("0" + value)
		for width in (12, 13, 14):
			variants.add(value.zfill(width))
			variants.add(stripped.zfill(width))
	return list(variants)


def _build_item_barcode_index():
	"""Map every barcode variant -> set of Item codes for bulk matching."""
	index = {}
	for row in frappe.db.sql(
		"select parent, barcode from `tabItem Barcode` where ifnull(barcode, '') != ''",
		as_dict=True,
	):
		for variant in _barcode_variants(row.barcode):
			index.setdefault(variant, set()).add(row.parent)
	return index


def _resolve_items_for_remote_barcodes(barcodes, barcode_index):
	"""Return (unique_item_codes, matched_variant) for one remote barcode list."""
	candidates = []
	matched_variant = None
	seen = set()
	for barcode in barcodes or []:
		for variant in _barcode_variants(barcode):
			if variant in seen:
				continue
			seen.add(variant)
			hits = barcode_index.get(variant) or set()
			if not hits:
				continue
			if matched_variant is None:
				matched_variant = variant
			candidates.extend(hits)
	# Some migrated rows use the barcode itself as Item.name.
	for barcode in barcodes or []:
		if barcode_index.get(barcode):
			continue
		if frappe.db.exists("Item", barcode):
			candidates.append(barcode)
			if matched_variant is None:
				matched_variant = barcode
	unique = list({code for code in candidates if code})
	return unique, matched_variant


def _catalog_sku(item_code, foodpanda_sku=None, barcode=None):
	"""Foodpanda catalog identity is barcode-based, never ERPNext Item Code.

	Mapped products already know Foodpanda's catalog SKU (often still a legacy
	vendorcode on Foodpanda's side after barcode match). Unmapped create/update
	attempts use the Item's primary barcode as the SKU.
	"""
	if foodpanda_sku:
		return str(foodpanda_sku).strip()
	barcode = barcode if barcode is not None else _get_barcode(item_code)
	if barcode:
		return str(barcode).strip()
	frappe.throw(
		_("Item {0} has no barcode; Foodpanda sync uses barcodes, not Item Codes").format(
			item_code
		)
	)


def _maximum_sales_quantity(quantity):
	"""Keep a customer order to one quarter of the branch's sellable stock.

	Foodpanda expects an integer. An in-stock item must allow at least one
	unit; zero stock remains inactive and receives a zero limit.
	"""
	if quantity <= 0:
		return 0
	return min(max(int(floor(quantity / 4)), 1), 36)


def _get_public_product_name(item_code, fallback):
	public_name = frappe.db.get_value("Shopping Product", {"item": item_code}, "public_name")
	return public_name or fallback


def build_update_payload(item_code, outlet, foodpanda_sku=None):
	"""Steady-state payload for an existing Foodpanda product - limited to
	the fields the docs' update example actually shows."""
	item = frappe.db.get_value(
		"Item", item_code, ["disabled", "is_sales_item"], as_dict=True
	)
	if not item:
		frappe.throw(_("Item {0} does not exist").format(item_code))

	price = _get_item_price(item_code, outlet.branch)
	quantity, in_stock = _get_stock(item_code, outlet.branch)
	# Partner API expects an integer quantity (float like 3.4 → HTTP 400).
	quantity = max(int(floor(quantity)), 0)
	active = bool(item.is_sales_item) and not item.disabled and quantity > 0
	if frappe.db.has_column("Foodpanda Product", "portal_active"):
		portal_active = frappe.db.get_value(
			"Foodpanda Product",
			{"item_code": item_code, "outlet": outlet.name},
			"portal_active",
		)
		# Explicit Catalog Sheet override: unchecked forces inactive on Foodpanda.
		if portal_active is not None and not int(portal_active):
			active = False
	barcode = _get_foodpanda_barcode(item_code)
	sku = _catalog_sku(item_code, foodpanda_sku=foodpanda_sku, barcode=barcode)

	payload = {
		"sku": sku,
		"active": active,
		"price": price,
		"quantity": quantity,
		"max_sales_quantity": _maximum_sales_quantity(quantity),
	}
	# Omit invalid local barcodes (e.g. SM1500, NC2) - Foodpanda returns HTTP 400
	# "barcode must contain only digits, and cannot be longer than 14 digits".
	if barcode:
		payload["barcode"] = barcode
	return payload


def build_create_payload(item_code, outlet):
	"""Documented Add Products payload with localized and array fields."""
	item = frappe.db.get_value(
		"Item", item_code, ["item_name", "description", "item_group", "disabled", "is_sales_item"], as_dict=True
	)
	if not item:
		frappe.throw(_("Item {0} does not exist").format(item_code))

	category_id = frappe.db.get_value("Foodpanda Category Map", item.item_group, "foodpanda_category_id")
	if not category_id:
		frappe.throw(_("Item Group {0} has no Foodpanda Category Map").format(item.item_group))

	barcode = _get_foodpanda_barcode(item_code)
	if not barcode:
		frappe.throw(
			_(
				"Item {0} needs a digit-only barcode (max 14 digits) to create a Foodpanda product"
			).format(item_code)
		)

	settings = client.get_settings()
	locale = settings.catalog_locale or "en_PK"
	payload = build_update_payload(item_code, outlet, foodpanda_sku=barcode)
	payload.pop("barcode", None)
	payload["title"] = {locale: _get_public_product_name(item_code, item.item_name)}
	payload["description"] = {locale: item.description or ""}
	payload["barcodes"] = [barcode]
	payload["categories"] = [category_id]
	return payload


def hash_payload(payload):
	return hashlib.sha256(json.dumps(payload, sort_keys=True).encode()).hexdigest()


def get_or_create_foodpanda_product(item_code, outlet_name):
	name = frappe.db.get_value("Foodpanda Product", {"item_code": item_code, "outlet": outlet_name}, "name")
	if name:
		return frappe.get_doc("Foodpanda Product", name)

	doc = frappe.get_doc(
		{
			"doctype": "Foodpanda Product",
			"item_code": item_code,
			"outlet": outlet_name,
			"sync_status": "Pending",
		}
	)
	doc.insert(ignore_permissions=True)
	return doc


def _submit_catalog_job(
	settings, method, chain_id, vendor_id, payload, outlet_name, include_api_response=False
):
	return _submit_catalog_products(
		settings,
		method,
		chain_id,
		vendor_id,
		[payload],
		outlet_name,
		include_api_response=include_api_response,
	)


def _submit_catalog_products(
	settings, method, chain_id, vendor_id, payloads, outlet_name, include_api_response=False
):
	"""Submit one or more catalog products in a single Partner API call."""
	if not payloads:
		raise FoodpandaAPIError("Foodpanda catalog submit called with no products")

	is_add = method == "POST"
	path = _ADD_PRODUCTS_PATH.format(chain_id=chain_id) if is_add else _UPDATE_PRODUCTS_PATH.format(
		chain_id=chain_id, vendor_id=vendor_id
	)
	body = {"vendors": [vendor_id], "products": payloads} if is_add else {"products": payloads}
	response = client.request(
		method,
		path,
		settings=settings,
		json=body,
	)
	response_data = response.json() or {}
	job_id = response_data.get("job_id")
	if not job_id:
		raise FoodpandaAPIError("Foodpanda catalog response had no job_id", response_body=response_data)

	frappe.get_doc(
		{
			"doctype": "Foodpanda Catalog Job",
			"job_id": job_id,
			"operation": "Add" if is_add else "Update",
			"outlet": outlet_name,
			"vendor_id": vendor_id,
			"status": "Pending",
			"requested_skus": json.dumps([row.get("sku") for row in payloads]),
			"request_payload": json.dumps(body, ensure_ascii=False, default=str),
			"raw_response": json.dumps(response_data, ensure_ascii=False, default=str),
			"submitted_at": now_datetime(),
		}
	).insert(ignore_permissions=True)
	if include_api_response:
		return job_id, {
			"http_status": response.status_code,
			"body": response_data,
		}
	return job_id




def sync_item(item_code, outlet_name):
	outlet = frappe.get_doc("Foodpanda Outlet", outlet_name)
	if not outlet.catalog_sync_enabled:
		return {"status": "Skipped", "reason": "Catalog sync disabled for this outlet"}

	product = get_or_create_foodpanda_product(item_code, outlet_name)
	is_new_product = not product.foodpanda_product_id
	settings = client.get_settings() if is_new_product else None
	if is_new_product and not settings.allow_product_creation:
		error_message = (
			"Product creation is disabled; map the existing Foodpanda barcode/SKU "
			"or enable beta product creation"
		)
		product.db_set({"sync_status": "Failed", "last_error": error_message})
		return {
			"status": "Failed",
			"error": error_message,
			"source": "ERPNext validation",
		}

	try:
		payload = (
			build_create_payload(item_code, outlet)
			if is_new_product
			else build_update_payload(
				item_code, outlet, foodpanda_sku=product.foodpanda_product_id
			)
		)
	except frappe.ValidationError as error:
		product.db_set({"sync_status": "Failed", "last_error": str(error)})
		return {"status": "Failed", "error": str(error), "source": "ERPNext validation"}

	content_hash = hash_payload(payload)

	if not is_new_product and product.sync_status == "Synced" and product.content_hash == content_hash:
		return {"status": "Synced", "skipped": True}
	settings = settings or client.get_settings()

	try:
		method = "POST" if is_new_product else "PUT"
		job_id, api_response = _submit_catalog_job(
			settings,
			method,
			settings.chain_id,
			outlet.vendor_id,
			payload,
			outlet_name,
			include_api_response=True,
		)
		product.db_set(
			{"sync_status": "Pending", "last_job_id": job_id, "pending_content_hash": content_hash, "last_error": ""}
		)
	except FoodpandaAPIError as error:
		client.log_api_failure(
			f"Foodpanda catalog sync failed: {item_code}", f"Outlet: {outlet_name}", error
		)
		product.db_set({"sync_status": "Failed", "last_error": str(error)})
		result = {"status": "Failed", "error": str(error), "source": "Foodpanda API"}
		if error.status_code is not None or error.response_body is not None:
			result["api_response"] = {
				"http_status": error.status_code,
				"body": error.response_body,
			}
		return result

	return {
		"status": "Pending",
		"job_id": job_id,
		"source": "Foodpanda API",
		"api_response": api_response,
	}


def sync_availability(item_code, branch):
	outlet_name = frappe.db.get_value(
		"Foodpanda Outlet", {"branch": branch, "catalog_sync_enabled": 1}, "name"
	)
	if not outlet_name:
		return

	product = frappe.db.get_value(
		"Foodpanda Product",
		{"item_code": item_code, "outlet": outlet_name},
		["name", "foodpanda_product_id"],
		as_dict=True,
	)
	# Nothing pushed yet for this item at this outlet - a full sync_item call
	# (bulk export or the next catalog change) carries stock/availability
	# with it, no need to push a partial update ahead of the item existing.
	if not product or not product.foodpanda_product_id:
		return

	outlet = frappe.get_doc("Foodpanda Outlet", outlet_name)
	settings = client.get_settings()

	try:
		payload = build_update_payload(
			item_code, outlet, foodpanda_sku=product.foodpanda_product_id
		)
		job_id = _submit_catalog_job(
			settings, "PUT", settings.chain_id, outlet.vendor_id, payload, outlet_name
		)
		frappe.db.set_value(
			"Foodpanda Product",
			product.name,
			{
				"sync_status": "Pending",
				"last_job_id": job_id,
				"pending_content_hash": hash_payload(payload),
				"last_error": "",
			},
		)
	except FoodpandaAPIError as error:
		client.log_api_failure(
			f"Foodpanda availability sync failed: {item_code}", f"Outlet: {outlet_name}", error
		)
		frappe.db.set_value("Foodpanda Product", product.name, {"sync_status": "Failed", "last_error": str(error)})
	except frappe.ValidationError as error:
		frappe.db.set_value("Foodpanda Product", product.name, {"sync_status": "Failed", "last_error": str(error)})


def iter_remote_catalog_products(outlet_name, page_size=500, source="auto"):
	"""Yield remote catalog products for one outlet.

	source:
	  - ``export``: latest Partner POST /catalog/export download only
	  - ``get``: paginated GET /catalog only
	  - ``auto`` (default): export file when available, else GET

	Foodpanda Partner API expects query param `page` (not `page_number`).
	Sending `page_number` is ignored and every request returns page 1, which
	looks like duplicate items / stuck pagination. Max `page_size` is 500.
	"""
	if source in ("auto", "export"):
		from aimatic.foodpanda_integration import catalog_export

		export_products = list(catalog_export.iter_cached_export_products(outlet_name))
		if export_products or source == "export":
			for product in export_products:
				yield product
			return

	if source not in ("auto", "get"):
		frappe.throw(_("Unknown Foodpanda catalog source: {0}").format(source))

	yield from _iter_remote_catalog_products_get(outlet_name, page_size=page_size)


def _iter_remote_catalog_products_get(outlet_name, page_size=500):
	"""Paginated GET /catalog — fallback when no export snapshot exists."""
	from aimatic.foodpanda_integration.catalog_export import normalize_remote_product

	outlet = frappe.get_doc("Foodpanda Outlet", outlet_name)
	settings = client.get_settings()
	path = _CATALOG_PATH.format(chain_id=settings.chain_id, vendor_id=outlet.vendor_id)
	page_size = min(max(int(page_size or 500), 1), 500)
	page = 1
	seen_skus = set()
	while True:
		response = client.request(
			"GET", path, settings=settings, params={"page": page, "page_size": page_size}
		)
		data = response.json() or {}
		products = data.get("products") or []
		if not products:
			break
		new_on_page = 0
		for product in products:
			sku = str(product.get("sku") or "").strip()
			if sku and sku in seen_skus:
				continue
			if sku:
				seen_skus.add(sku)
			new_on_page += 1
			normalized = normalize_remote_product(product) or product
			yield normalized
		total_pages = int(data.get("total_pages") or page)
		# Stop when API says we're past the last page, or a page added nothing
		# new (guards against a stuck cursor still returning page 1).
		if page >= total_pages or new_on_page == 0:
			break
		page += 1


def _catalog_source_label(source):
	return {
		"export": "Partner catalog export download",
		"get": "Partner GET /catalog pagination",
		"auto": "Partner catalog export (fallback GET /catalog)",
	}[source]


def map_remote_catalog_by_barcode(outlet_name, source="auto"):
	"""Match Foodpanda catalog rows to Items by barcode and store Foodpanda SKU.

	Update-only scope: does not create products on Foodpanda. Sets
	Foodpanda Product.foodpanda_product_id to the remote catalog SKU so later
	PUT calls address the existing Foodpanda row. Matching is barcode-only
	(including Foodpanda's extra leading-zero padding). Writes an Excel report
	attached to the outlet.

	Prefer ``source='export'`` or ``'auto'`` after POST /catalog/export so the
	full store assortment is read from one download instead of many GET pages.
	"""
	from aimatic.foodpanda_integration import catalog_export

	outlet = frappe.get_doc("Foodpanda Outlet", outlet_name)
	if not outlet.catalog_sync_enabled:
		frappe.throw(_("Catalog sync is not enabled for this outlet"))

	if source in ("auto", "export") and not catalog_export.cached_export_product_count(outlet_name):
		if source == "export":
			frappe.throw(
				_(
					"No completed catalog export for this outlet. Use Import Catalog from Foodpanda first."
				)
			)
		source = "get"

	barcode_index = _build_item_barcode_index()
	report_rows = []
	mapped = 0
	updated = 0
	skipped_no_barcode = 0
	skipped_unmatched = 0
	skipped_ambiguous = 0
	remote_total = 0
	resolved_source = "export" if source == "auto" and catalog_export.cached_export_product_count(outlet_name) else source

	for remote in iter_remote_catalog_products(outlet_name, source=resolved_source):
		remote_total += 1
		remote_sku = str(remote.get("sku") or "").strip()
		barcodes = remote.get("barcodes") or []
		if not isinstance(barcodes, list):
			barcodes = [barcodes] if barcodes else []
		barcodes = [str(b).strip() for b in barcodes if b]
		title = remote.get("title") or ""
		price = remote.get("price")
		active = remote.get("active")
		variants_tried = _variants_tried_for_barcodes(barcodes)
		base_row = {
			"foodpanda_sku": remote_sku,
			"foodpanda_barcodes": ", ".join(barcodes),
			"variants_tried": variants_tried,
			"foodpanda_title": title,
			"foodpanda_price": price,
			"foodpanda_active": active,
			"item_code": "",
			"item_name": "",
			"matched_barcode": "",
			"erpnext_barcodes": "",
			"notes": "",
		}

		if not remote_sku:
			skipped_unmatched += 1
			report_rows.append({**base_row, "status": "Missing SKU", "notes": "Remote catalog row has no SKU"})
			continue
		if not barcodes:
			skipped_no_barcode += 1
			report_rows.append({
				**base_row,
				"status": "No Barcode",
				"notes": "Partner API returned no barcodes for this Foodpanda product; cannot match to ERPNext",
			})
			continue

		unique_items, matched_variant = _resolve_items_for_remote_barcodes(barcodes, barcode_index)
		if not unique_items:
			skipped_unmatched += 1
			report_rows.append({
				**base_row,
				"status": "Unmatched",
				"notes": "No ERPNext Item Barcode matched these variants",
			})
			continue
		if len(unique_items) > 1:
			skipped_ambiguous += 1
			report_rows.append({
				**base_row,
				"status": "Ambiguous",
				"item_code": ", ".join(unique_items),
				"matched_barcode": matched_variant or "",
				"erpnext_barcodes": "; ".join(
					f"{code}: {_erpnext_barcodes_for_item(code)}" for code in unique_items
				),
				"notes": "Multiple Items share this barcode",
			})
			continue

		item_code = unique_items[0]
		item_name = frappe.db.get_value("Item", item_code, "item_name") or ""
		erpnext_barcodes = _erpnext_barcodes_for_item(item_code)
		product = get_or_create_foodpanda_product(item_code, outlet_name)
		if product.foodpanda_product_id == remote_sku:
			mapped += 1
			report_rows.append({
				**base_row,
				"status": "Mapped",
				"item_code": item_code,
				"item_name": item_name,
				"matched_barcode": matched_variant or "",
				"erpnext_barcodes": erpnext_barcodes,
				"notes": "Already mapped",
			})
			continue
		product.db_set(
			{
				"foodpanda_product_id": remote_sku,
				"sync_status": "Pending",
				"last_error": "",
			}
		)
		mapped += 1
		updated += 1
		report_rows.append({
			**base_row,
			"status": "Updated",
			"item_code": item_code,
			"item_name": item_name,
			"matched_barcode": matched_variant or "",
			"erpnext_barcodes": erpnext_barcodes,
			"notes": "Foodpanda Product ID set from remote SKU",
		})

	local_rows = _local_barcode_report_rows()
	mapped_count = frappe.db.count(
		"Foodpanda Product",
		{"outlet": outlet_name, "foodpanda_product_id": ("is", "set")},
	)
	outlet.db_set(
		{
			"remote_sku_count": remote_total,
			"mapped_sku_count": mapped_count,
		}
	)
	summary = {
		"outlet": outlet_name,
		"catalog_source": _catalog_source_label(resolved_source),
		"remote_total": remote_total,
		"mapped": mapped,
		"updated": updated,
		"skipped_no_barcode": skipped_no_barcode,
		"skipped_unmatched": skipped_unmatched,
		"skipped_ambiguous": skipped_ambiguous,
		"erpnext_barcode_rows": len(local_rows),
		"foodpanda_products_mapped_total": mapped_count,
		"note": (
			"Matching uses barcodes only (never ERPNext Item Code). "
			"Import Catalog from Foodpanda (POST /catalog/export) is recommended "
			"for the full store assortment; GET /catalog pagination is the fallback."
		),
	}
	file_url = _save_matching_report(
		outlet_name,
		_matching_report_bytes(report_rows, local_barcode_rows=local_rows, summary=summary),
	)
	return {
		"outlet": outlet_name,
		"catalog_source": resolved_source,
		"remote_total": remote_total,
		"mapped": mapped,
		"updated": updated,
		"skipped_no_barcode": skipped_no_barcode,
		"skipped_unmatched": skipped_unmatched,
		"skipped_ambiguous": skipped_ambiguous,
		"mapped_total": mapped_count,
		"file_url": file_url,
	}


def start_catalog_import(outlet_name):
	"""Request a Partner catalog export (POST /catalog/export)."""
	outlet = frappe.get_doc("Foodpanda Outlet", outlet_name)
	if not outlet.catalog_sync_enabled:
		frappe.throw(_("Catalog sync is not enabled for this outlet"))
	return request_remote_export(outlet_name)


def start_import_and_map(outlet_name):
	"""Background: export download then barcode mapping."""
	outlet = frappe.get_doc("Foodpanda Outlet", outlet_name)
	if not outlet.catalog_sync_enabled:
		frappe.throw(_("Catalog sync is not enabled for this outlet"))

	result = request_remote_export(outlet_name)
	frappe.enqueue(
		"aimatic.foodpanda_integration.catalog.run_import_and_map",
		queue="long",
		outlet_name=outlet_name,
		foodpanda_job_id=result["job_id"],
		enqueue_after_commit=True,
		timeout=7200,
		job_name=f"Foodpanda import and map {outlet_name}",
	)
	return result


def run_import_and_map(outlet_name, foodpanda_job_id):
	from aimatic.foodpanda_integration import catalog_export

	catalog_export.wait_for_export_ready(foodpanda_job_id)
	return map_remote_catalog_by_barcode(outlet_name, source="export")


def start_bulk_push(outlet_name, item_codes=None):
	"""Enqueue bulk PUT of price/stock for mapped Foodpanda Products."""
	return start_bulk_export(outlet_name, item_codes=item_codes)


def _matching_report_bytes(rows, local_barcode_rows=None, summary=None):
	from openpyxl import Workbook

	workbook = Workbook()
	sheet = workbook.active
	sheet.title = "Foodpanda Matching"
	headers = [
		"Match Status",
		"Foodpanda SKU",
		"Foodpanda Barcodes",
		"Variants Tried",
		"Foodpanda Title",
		"Foodpanda Price",
		"Foodpanda Active",
		"Matched Item Code",
		"Matched Item Name",
		"Matched On Barcode",
		"ERPNext Barcodes On Item",
		"Notes",
	]
	sheet.append(headers)
	for row in rows:
		sheet.append([row.get(key) for key in [
			"status",
			"foodpanda_sku",
			"foodpanda_barcodes",
			"variants_tried",
			"foodpanda_title",
			"foodpanda_price",
			"foodpanda_active",
			"item_code",
			"item_name",
			"matched_barcode",
			"erpnext_barcodes",
			"notes",
		]])

	local_sheet = workbook.create_sheet("ERPNext Barcodes")
	local_sheet.append(["ERPNext Barcode", "Item Code", "Item Name"])
	for row in local_barcode_rows or []:
		local_sheet.append([row.get("barcode"), row.get("item_code"), row.get("item_name")])

	summary_sheet = workbook.create_sheet("Summary")
	summary_sheet.append(["Field", "Value"])
	for key, value in (summary or {}).items():
		summary_sheet.append([key, value])

	buffer = io.BytesIO()
	workbook.save(buffer)
	return buffer.getvalue()


def _local_barcode_report_rows():
	"""One row per stored Item Barcode (not expanded variants)."""
	return frappe.db.sql(
		"""
		select ib.barcode, ib.parent as item_code, ifnull(i.item_name, '') as item_name
		from `tabItem Barcode` ib
		left join `tabItem` i on i.name = ib.parent
		where ifnull(ib.barcode, '') != ''
		order by ib.barcode
		""",
		as_dict=True,
	)


def _erpnext_barcodes_for_item(item_code):
	return ", ".join(
		frappe.get_all("Item Barcode", filters={"parent": item_code}, pluck="barcode") or []
	)


def _variants_tried_for_barcodes(barcodes):
	seen = []
	for barcode in barcodes or []:
		for variant in _barcode_variants(barcode):
			if variant not in seen:
				seen.append(variant)
	return ", ".join(seen)


def _save_matching_report(outlet_name, content):
	from frappe.utils.file_manager import save_file

	safe_outlet = "".join(ch if ch.isalnum() or ch in "-_" else "_" for ch in outlet_name)[:40]
	filename = f"foodpanda-barcode-matching-{safe_outlet}-{frappe.utils.now_datetime().strftime('%Y%m%d-%H%M%S')}.xlsx"
	saved = save_file(
		filename,
		content,
		"Foodpanda Outlet",
		outlet_name,
		is_private=1,
	)
	return saved.file_url


def get_remote_catalog(outlet_name):
	outlet = frappe.get_doc("Foodpanda Outlet", outlet_name)
	settings = client.get_settings()
	response = client.request(
		"GET",
		_CATALOG_PATH.format(chain_id=settings.chain_id, vendor_id=outlet.vendor_id),
		settings=settings,
	)
	return response.json()


def get_remote_categories(outlet_name):
	outlet = frappe.get_doc("Foodpanda Outlet", outlet_name)
	settings = client.get_settings()
	response = client.request(
		"GET",
		_CATEGORIES_PATH.format(chain_id=settings.chain_id, vendor_id=outlet.vendor_id),
		settings=settings,
	)
	return response.json()


def request_remote_export(outlet_name):
	outlet = frappe.get_doc("Foodpanda Outlet", outlet_name)
	settings = client.get_settings()
	response = client.request(
		"POST",
		_EXPORT_PATH.format(chain_id=settings.chain_id, vendor_id=outlet.vendor_id),
		settings=settings,
	)
	response_data = response.json() or {}
	job_id = response_data.get("job_id")
	if not job_id:
		raise FoodpandaAPIError("Foodpanda catalog export response had no job_id", response_body=response_data)
	frappe.get_doc({
		"doctype": "Foodpanda Catalog Job",
		"job_id": job_id,
		"operation": "Export",
		"outlet": outlet_name,
		"vendor_id": outlet.vendor_id,
		"status": "Pending",
		"raw_response": json.dumps(response_data, ensure_ascii=False, default=str),
		"submitted_at": now_datetime(),
	}).insert(ignore_permissions=True)
	return {"job_id": job_id, "status": "Pending"}


def refresh_remote_job(job_id):
	from aimatic.foodpanda_integration import catalog_jobs

	settings = client.get_settings()
	response = client.request(
		"GET", _JOB_STATUS_PATH.format(chain_id=settings.chain_id, job_id=job_id), settings=settings
	)
	payload = response.json() or {}
	payload.setdefault("job_id", job_id)
	return catalog_jobs.process_callback(payload)


def _job_cache_key(job_id):
	return f"{_JOB_CACHE_PREFIX}{job_id}"


def _set_job_status(job_id, values):
	current = _get_job_status(job_id) or {}
	current.update(values)
	frappe.cache.set_value(_job_cache_key(job_id), json.dumps(current), expires_in_sec=_JOB_CACHE_TTL)


def _get_job_status(job_id):
	raw = frappe.cache.get_value(_job_cache_key(job_id))
	if not raw:
		return None
	if isinstance(raw, bytes):
		raw = raw.decode()
	return json.loads(raw)



def start_bulk_export(outlet_name, item_codes=None):
	outlet = frappe.get_doc("Foodpanda Outlet", outlet_name)
	if not outlet.catalog_sync_enabled:
		frappe.throw(_("Catalog sync is not enabled for this outlet"))

	if isinstance(item_codes, str):
		item_codes = frappe.parse_json(item_codes)
	if item_codes is not None and not isinstance(item_codes, (list, tuple)):
		frappe.throw(_("item_codes must be a list of Item Codes"))
	item_codes = [str(code).strip() for code in (item_codes or []) if str(code).strip()] or None

	export_job_id = uuid.uuid4().hex
	_set_job_status(
		export_job_id,
		{
			"job_id": export_job_id,
			"outlet": outlet_name,
			"owner": frappe.session.user,
			"status": "queued",
			"synced": 0,
			"failed": 0,
			"total": 0,
			"item_filter": len(item_codes) if item_codes else 0,
		},
	)
	# Do not pass `job_id=` into enqueue - Frappe/RQ reserves that for the
	# Redis job identity and will not forward it to run_bulk_export.
	frappe.enqueue(
		"aimatic.foodpanda_integration.catalog.run_bulk_export",
		queue="long",
		export_job_id=export_job_id,
		outlet_name=outlet_name,
		item_codes=item_codes,
		enqueue_after_commit=True,
		timeout=7200,
		job_name=f"Foodpanda bulk export {outlet_name}",
	)
	return _get_job_status(export_job_id)


def get_bulk_export_status(job_id):
	status = _get_job_status(job_id)
	if not status:
		frappe.throw(_("Export job expired or was not found"))
	if status.get("owner") != frappe.session.user:
		frappe.throw(_("This export job is not available to you"), frappe.PermissionError)
	return status


_BULK_UPDATE_BATCH_SIZE = 50
_BULK_UPDATE_PAUSE_SECONDS = 1.0


def run_bulk_export(export_job_id=None, outlet_name=None, job_id=None, item_codes=None):
	"""Background worker for Push Price & Stock (bulk PUT).

	Mapped products are sent in Partner API bulk PUT batches (not one HTTP
	call per item) to stay within Foodpanda rate limits.
	"""
	import time

	export_job_id = export_job_id or job_id
	if not export_job_id or not outlet_name:
		frappe.log_error(
			title="Foodpanda bulk export missing arguments",
			message=f"export_job_id={export_job_id!r} outlet_name={outlet_name!r}",
		)
		return

	status = _get_job_status(export_job_id)
	if not status:
		return

	outlet = frappe.get_doc("Foodpanda Outlet", outlet_name)
	if not outlet.catalog_sync_enabled:
		_set_job_status(export_job_id, {"status": "done", "failed": 0, "synced": 0, "total": 0})
		return

	if isinstance(item_codes, str):
		item_codes = frappe.parse_json(item_codes)
	item_codes = [str(code).strip() for code in (item_codes or []) if str(code).strip()]

	if item_codes:
		rows = frappe.db.sql(
			"""
			select name, item_code, foodpanda_product_id, sync_status, content_hash
			from `tabFoodpanda Product`
			where outlet = %s
				and ifnull(foodpanda_product_id, '') != ''
				and item_code in %s
			order by item_code
			""",
			(outlet_name, tuple(item_codes)),
			as_dict=True,
		)
	else:
		rows = frappe.db.sql(
			"""
			select name, item_code, foodpanda_product_id, sync_status, content_hash
			from `tabFoodpanda Product`
			where outlet = %s and ifnull(foodpanda_product_id, '') != ''
			order by item_code
			""",
			outlet_name,
			as_dict=True,
		)
	_set_job_status(export_job_id, {"status": "running", "total": len(rows)})

	settings = client.get_settings()
	synced, failed = 0, 0
	batch = []

	def flush_batch():
		nonlocal synced, failed, batch
		if not batch:
			return
		payloads = [row["payload"] for row in batch]
		try:
			fp_job_id = _submit_catalog_products(
				settings,
				"PUT",
				settings.chain_id,
				outlet.vendor_id,
				payloads,
				outlet_name,
			)
			for row in batch:
				frappe.db.set_value(
					"Foodpanda Product",
					row["name"],
					{
						"sync_status": "Pending",
						"last_job_id": fp_job_id,
						"pending_content_hash": row["content_hash"],
						"last_error": "",
					},
				)
			synced += len(batch)
		except FoodpandaAPIError as error:
			client.log_api_failure(
				f"Foodpanda bulk catalog sync failed ({len(batch)} products)",
				f"Outlet: {outlet_name}",
				error,
			)
			for row in batch:
				frappe.db.set_value(
					"Foodpanda Product",
					row["name"],
					{"sync_status": "Failed", "last_error": str(error)},
				)
			failed += len(batch)
		except Exception:
			frappe.log_error(
				frappe.get_traceback(),
				f"Foodpanda bulk export batch failed ({len(batch)} products)",
			)
			for row in batch:
				frappe.db.set_value(
					"Foodpanda Product",
					row["name"],
					{"sync_status": "Failed", "last_error": "Bulk export batch failed"},
				)
			failed += len(batch)
		_set_job_status(export_job_id, {"synced": synced, "failed": failed})
		frappe.db.commit()
		batch = []
		time.sleep(_BULK_UPDATE_PAUSE_SECONDS)

	for row in rows:
		try:
			payload = build_update_payload(
				row.item_code, outlet, foodpanda_sku=row.foodpanda_product_id
			)
			content_hash = hash_payload(payload)
			if row.sync_status == "Synced" and row.content_hash == content_hash:
				synced += 1
				_set_job_status(export_job_id, {"synced": synced, "failed": failed})
				continue
			batch.append(
				{
					"name": row.name,
					"item_code": row.item_code,
					"payload": payload,
					"content_hash": content_hash,
				}
			)
			if len(batch) >= _BULK_UPDATE_BATCH_SIZE:
				flush_batch()
		except Exception as error:
			frappe.db.set_value(
				"Foodpanda Product",
				row.name,
				{"sync_status": "Failed", "last_error": str(error)[:1400]},
			)
			failed += 1
			_set_job_status(export_job_id, {"synced": synced, "failed": failed})

	flush_batch()
	_set_job_status(export_job_id, {"status": "done"})
