import collections

import frappe
import openpyxl

FILE_PATH = "/home/nabeel/frappe-bench/sites/szl/public/files/itemmastersto.xlsx"
SHEET_NAME = "Query1"
MIN_CODE_LEN = 8
COMMIT_EVERY = 2000
DRY_RUN = False


def get_barcode_to_items():
    """barcode -> list of Item names, since one barcode can legitimately be
    reused across more than one Item and every match must be updated."""
    mapping = collections.defaultdict(list)
    for row in frappe.get_all("Item Barcode", fields=["barcode", "parent"]):
        if row.barcode:
            mapping[row.barcode].append(row.parent)
    return mapping


def ensure_brand(brand_name, existing_brands, stats):
    if brand_name in existing_brands:
        return
    if DRY_RUN:
        stats["would_create_brands"].add(brand_name)
        return
    if not frappe.db.exists("Brand", brand_name):
        frappe.get_doc({"doctype": "Brand", "brand": brand_name}).insert(ignore_permissions=True)
    existing_brands.add(brand_name)


def iter_rows():
    wb = openpyxl.load_workbook(FILE_PATH, read_only=True, data_only=True)
    ws = wb[SHEET_NAME]
    header = None
    for row in ws.iter_rows(values_only=True):
        if header is None:
            header = row
            continue
        if not any(row):
            continue
        yield dict(zip(header, row))


def run():
    barcode_to_items = get_barcode_to_items()
    existing_brands = set(frappe.get_all("Brand", pluck="name"))

    counts = collections.Counter()
    stats = {"would_create_brands": set()}
    unmatched_samples = []
    updates_since_commit = 0

    for index, row in enumerate(iter_rows(), start=1):
        counts["rows_seen"] += 1

        item_code_raw = str(row.get("ItemCode") or "").strip()
        if len(item_code_raw) < MIN_CODE_LEN:
            counts["skipped_short_code"] += 1
            continue

        brand = str(row.get("AiBrand") or "").strip()
        if not brand or brand.lower() == "generic":
            counts["skipped_generic_brand"] += 1
            continue

        matched_items = barcode_to_items.get(item_code_raw)
        if not matched_items:
            counts["skipped_unmatched_barcode"] += 1
            if len(unmatched_samples) < 50:
                unmatched_samples.append(item_code_raw)
            continue

        if len(matched_items) > 1:
            counts["multi_match_barcodes"] += 1

        ensure_brand(brand, existing_brands, stats)

        for item_name in matched_items:
            if not DRY_RUN:
                frappe.db.set_value("Item", item_name, "brand", brand, update_modified=False)
            counts["items_updated" if not DRY_RUN else "items_would_update"] += 1
            updates_since_commit += 1

        if not DRY_RUN and updates_since_commit >= COMMIT_EVERY:
            frappe.db.commit()
            updates_since_commit = 0

        if index % 20000 == 0:
            print(f"Processed {index} rows...")

    if not DRY_RUN:
        frappe.db.commit()

    print("DRY_RUN", DRY_RUN)
    print("SUMMARY", dict(counts))
    print("NEW_BRANDS_COUNT", len(stats["would_create_brands"]))
    print("UNMATCHED_SAMPLE", unmatched_samples)


run()
