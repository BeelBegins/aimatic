import collections

import frappe
import openpyxl

FILE_PATH = "/home/nabeel/frappe-bench/sites/szl/public/files/itemmastersto.xlsx"
SHEET_NAME = "Query1"
MIN_CODE_LEN = 8
COMMIT_EVERY = 2000
DRY_RUN = False

# The sheet mislabels several unrelated FMCG Food subcategories (Dry Fruits &
# Nuts, Pulses) under a category literally named "Arfa Food" -- looks like a
# supplier/brand name that leaked into what should be a generic category
# label. Normalized to a real generic category name before the tree is built.
CATEGORY_RENAMES = {
	"arfa food": "Dry Foods",
}


def get_barcode_to_items():
	mapping = collections.defaultdict(list)
	for row in frappe.get_all("Item Barcode", fields=["barcode", "parent"]):
		if row.barcode:
			mapping[row.barcode].append(row.parent)
	return mapping


def iter_rows():
	wb = openpyxl.load_workbook(FILE_PATH, read_only=True, data_only=True)
	ws = wb[SHEET_NAME]
	header = None
	for row in ws.iter_rows(values_only=True):
		if header is None:
			header = row
			continue
		if not any(row):
			continue
		yield dict(zip(header, row))


def normalize_category(cat):
	return CATEGORY_RENAMES.get(cat.lower(), cat)


def resolve_item_categories():
	"""item_name -> resolved (dept, cat), using most-frequent-pair-wins with
	last-row-as-tiebreak across that item's matched sheet rows."""
	barcode_to_items = get_barcode_to_items()

	# item_name -> {(dept, cat): [count, last_seen_index]}
	candidates = collections.defaultdict(lambda: collections.defaultdict(lambda: [0, -1]))
	rows_seen = 0
	rows_matched = 0
	rows_missing_category = 0

	for index, row in enumerate(iter_rows()):
		item_code_raw = str(row.get("ItemCode") or "").strip()
		if len(item_code_raw) < MIN_CODE_LEN:
			continue
		matched = barcode_to_items.get(item_code_raw)
		if not matched:
			continue
		rows_seen += 1

		dept = str(row.get("AiDepartment") or "").strip()
		cat = str(row.get("AiCategory") or "").strip()
		if not dept or not cat:
			rows_missing_category += 1
			continue
		cat = normalize_category(cat)
		rows_matched += 1

		for item_name in matched:
			entry = candidates[item_name][(dept, cat)]
			entry[0] += 1
			entry[1] = index

	resolved = {}
	conflicts_resolved = 0
	for item_name, pair_counts in candidates.items():
		if len(pair_counts) > 1:
			conflicts_resolved += 1
		best_pair = max(pair_counts.items(), key=lambda kv: (kv[1][0], kv[1][1]))[0]
		resolved[item_name] = best_pair

	meta = {
		"rows_seen": rows_seen,
		"rows_missing_category": rows_missing_category,
		"rows_used": rows_matched,
		"distinct_items_resolved": len(resolved),
		"conflicts_resolved": conflicts_resolved,
	}
	return resolved, meta


def ensure_item_group(name, parent, is_group, created_counter):
	if frappe.db.exists("Item Group", name):
		return
	if DRY_RUN:
		created_counter.add(name)
		return
	frappe.get_doc(
		{
			"doctype": "Item Group",
			"item_group_name": name,
			"parent_item_group": parent,
			"is_group": is_group,
		}
	).insert(ignore_permissions=True)
	created_counter.add(name)


def run():
	resolved, meta = resolve_item_categories()

	departments_created = set()
	categories_created = set()
	items_updated = 0
	updates_since_commit = 0

	for item_name, (dept, cat) in resolved.items():
		ensure_item_group(dept, "All Item Groups", 1, departments_created)
		ensure_item_group(cat, dept, 0, categories_created)

		if not DRY_RUN:
			frappe.db.set_value("Item", item_name, "item_group", cat, update_modified=False)
		items_updated += 1
		updates_since_commit += 1

		if not DRY_RUN and updates_since_commit >= COMMIT_EVERY:
			frappe.db.commit()
			updates_since_commit = 0

	if not DRY_RUN:
		frappe.db.commit()

	print("DRY_RUN", DRY_RUN)
	print("META", meta)
	print("DEPARTMENTS", len(departments_created), sorted(departments_created))
	print("CATEGORIES", len(categories_created))
	print("ITEMS_UPDATED" if not DRY_RUN else "ITEMS_WOULD_UPDATE", items_updated)


run()
