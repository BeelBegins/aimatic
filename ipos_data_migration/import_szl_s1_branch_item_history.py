"""Import branch past Sale/Debit/Credit rows into Branch Item History.

Source shape (Sheet1):
  ItemCode, Description, CostTotal, SaleTotal, Quantity, Date, Supcode,
  TransactionType (Sale|Debit|Credit), Store

Does NOT post Stock Ledger, GL, or Purchase documents — shadow history only.

Run via bench console (pass globals explicitly):
    exec(open("apps/aimatic/ipos_data_migration/import_szl_s1_branch_item_history.py").read(), globals())
"""

from __future__ import annotations

import hashlib
import os
from collections import Counter
from datetime import datetime

import frappe
from frappe.utils import cint, flt, getdate, now_datetime
from openpyxl import load_workbook

# --- Edit these before each run -------------------------------------------------
TARGET_SITE = "szl"
BRANCH = "S1 - Ghouri Town VIP"
COMPANY = "Siezal Supermarket"
FILE_PATH = "/home/nabeel/frappe-bench/sites/szl/private/files/historys1.xlsx"
IMPORT_TAG = "S1-HISTORY-2026-08-23"
# True -> parse + match stats only; no delete/insert/commit.
DRY_RUN = True
# When False and IMPORT_TAG already has rows, abort. When True, delete that tag first.
REPLACE_EXISTING_TAG = False
CHUNK_SIZE = 2000
# ---------------------------------------------------------------------------------

REQUIRED_HEADERS = {
	"ItemCode",
	"Description",
	"CostTotal",
	"SaleTotal",
	"Quantity",
	"Date",
	"Supcode",
	"TransactionType",
	"Store",
}
VALID_TYPES = {"Sale", "Debit", "Credit"}


def _guard():
	site = frappe.local.site
	if site != TARGET_SITE and not site.startswith(TARGET_SITE + "."):
		frappe.throw(f"Refusing to run on site {site!r}; expected {TARGET_SITE!r}")
	if not frappe.db.exists("Branch", BRANCH):
		frappe.throw(f"Branch {BRANCH!r} not found")
	if not frappe.db.exists("Company", COMPANY):
		frappe.throw(f"Company {COMPANY!r} not found")
	if not os.path.isfile(FILE_PATH):
		frappe.throw(f"Source file not found: {FILE_PATH}")
	if not frappe.db.exists("DocType", "Branch Item History"):
		frappe.throw("DocType 'Branch Item History' missing — run bench migrate first")


def _as_float(value) -> float:
	if value is None or value == "":
		return 0.0
	return flt(value)


def _as_date(value):
	if value is None or value == "":
		return None
	if isinstance(value, datetime):
		return getdate(value)
	return getdate(value)


def _fingerprint(parts: list[str]) -> str:
	raw = "|".join(parts)
	return hashlib.sha1(raw.encode("utf-8")).hexdigest()


def _build_item_maps():
	item_codes = set(frappe.get_all("Item", pluck="name"))
	barcode_to_item = {
		row.barcode: row.parent
		for row in frappe.get_all("Item Barcode", fields=["barcode", "parent"])
		if row.barcode
	}
	return item_codes, barcode_to_item


def _resolve_item(raw: str, item_codes: set[str], barcode_to_item: dict[str, str]):
	code = (raw or "").strip()
	if not code:
		return None, "Unmatched"
	if code in item_codes:
		return code, "Matched Item"
	item = barcode_to_item.get(code)
	if item:
		return item, "Matched Barcode"
	return None, "Unmatched"


def _iter_excel_rows(path: str):
	wb = load_workbook(path, read_only=True, data_only=True)
	try:
		ws = wb[wb.sheetnames[0]]
		header = None
		for excel_row_no, row in enumerate(ws.iter_rows(values_only=True), start=1):
			if header is None:
				header = [str(c).strip() if c is not None else "" for c in row]
				missing = REQUIRED_HEADERS - set(header)
				if missing:
					frappe.throw(f"Source sheet missing columns: {sorted(missing)}")
				continue
			if not any(v is not None and str(v).strip() for v in row):
				continue
			record = {header[i]: row[i] if i < len(row) else None for i in range(len(header))}
			yield excel_row_no, record
	finally:
		wb.close()


def _existing_tag_count() -> int:
	return cint(
		frappe.db.count("Branch Item History", {"import_tag": IMPORT_TAG, "branch": BRANCH})
	)


def run():
	_guard()
	source_file = os.path.basename(FILE_PATH)
	item_codes, barcode_to_item = _build_item_maps()

	stats = Counter()
	rows_out: list[tuple] = []
	now = now_datetime()
	user = frappe.session.user or "Administrator"

	fields = [
		"name",
		"creation",
		"modified",
		"modified_by",
		"owner",
		"docstatus",
		"idx",
		"branch",
		"company",
		"posting_date",
		"transaction_type",
		"item_code",
		"item_code_raw",
		"item_name",
		"match_status",
		"qty",
		"cost_total",
		"sale_total",
		"legacy_supcode",
		"store_name",
		"source_file",
		"source_row",
		"import_tag",
		"row_fingerprint",
	]

	for source_row, record in _iter_excel_rows(FILE_PATH):
		stats["source_rows"] += 1
		raw_item = str(record.get("ItemCode") or "").strip()
		ttype = str(record.get("TransactionType") or "").strip()
		posting_date = _as_date(record.get("Date"))
		qty = _as_float(record.get("Quantity"))
		cost_total = _as_float(record.get("CostTotal"))
		sale_total = _as_float(record.get("SaleTotal"))
		supcode = record.get("Supcode")
		legacy_supcode = "" if supcode is None else str(supcode).strip()
		store_name = str(record.get("Store") or "").strip()
		item_name = str(record.get("Description") or "").strip()

		if ttype not in VALID_TYPES:
			stats["skipped_bad_type"] += 1
			continue
		if not raw_item:
			stats["skipped_no_item"] += 1
			continue
		if not posting_date:
			stats["skipped_no_date"] += 1
			continue

		item_code, match_status = _resolve_item(raw_item, item_codes, barcode_to_item)
		stats[match_status] += 1
		stats[f"type_{ttype}"] += 1

		fp = _fingerprint(
			[
				IMPORT_TAG,
				BRANCH,
				str(source_row),
				raw_item,
				str(posting_date),
				ttype,
				f"{qty:.6f}",
				f"{cost_total:.6f}",
				f"{sale_total:.6f}",
				legacy_supcode,
			]
		)
		rows_out.append(
			(
				fp,
				now,
				now,
				user,
				user,
				0,
				0,
				BRANCH,
				COMPANY,
				posting_date,
				ttype,
				item_code,
				raw_item,
				item_name[:140] if item_name else None,
				match_status,
				qty,
				cost_total,
				sale_total,
				legacy_supcode or None,
				store_name or None,
				source_file,
				source_row,
				IMPORT_TAG,
				fp,
			)
		)

	existing = _existing_tag_count()
	print("IMPORT SUMMARY (pre-write)", dict(stats))
	print(
		{
			"branch": BRANCH,
			"company": COMPANY,
			"file": FILE_PATH,
			"import_tag": IMPORT_TAG,
			"ready_rows": len(rows_out),
			"existing_tag_rows": existing,
			"dry_run": DRY_RUN,
			"replace_existing_tag": REPLACE_EXISTING_TAG,
		}
	)

	if DRY_RUN:
		print("DRY_RUN=True — no DB writes")
		return dict(stats)

	if existing and not REPLACE_EXISTING_TAG:
		frappe.throw(
			f"Import tag {IMPORT_TAG!r} already has {existing} rows for {BRANCH}. "
			"Set REPLACE_EXISTING_TAG=True to delete and re-import."
		)

	if existing and REPLACE_EXISTING_TAG:
		frappe.db.sql(
			"""
			DELETE FROM `tabBranch Item History`
			WHERE import_tag = %s AND branch = %s
			""",
			(IMPORT_TAG, BRANCH),
		)
		print(f"Deleted {existing} existing rows for tag {IMPORT_TAG}")

	frappe.db.bulk_insert("Branch Item History", fields=fields, values=rows_out, chunk_size=CHUNK_SIZE)
	frappe.db.commit()
	print(f"Inserted {len(rows_out)} Branch Item History rows")
	return dict(stats)


# Auto-run when exec()'d from bench console.
run()
