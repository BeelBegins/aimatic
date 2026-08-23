"""Assign SIEZAL Item.brand from itemmastersto.xlsx by Item Barcode.

Run from the bench root with:
    echo 'exec(open("apps/aimatic/ipos_data_migration/update_siezal_item_brands.py").read(), globals())' \
        | bench --site siezal console

Keep DRY_RUN enabled first. The source ItemCode is a barcode, not ERPNext's
Item.item_code. Blank/Generic source brands and short/junk codes are ignored.
"""

import collections

import frappe
import openpyxl

TARGET_SITE = "siezal"
FILE_PATH = "/home/nabeel/frappe-bench/sites/szl/public/files/itemmastersto.xlsx"
SHEET_NAME = "Query1"
MIN_CODE_LEN = 8
COMMIT_EVERY = 2000
DRY_RUN = True


def get_barcode_to_items():
	"""Return barcode -> Item names; update every Item if reuse ever exists."""
	mapping = collections.defaultdict(list)
	for row in frappe.get_all("Item Barcode", fields=["barcode", "parent"]):
		if row.barcode:
			mapping[str(row.barcode).strip()].append(row.parent)
	return mapping


def iter_rows():
	workbook = openpyxl.load_workbook(FILE_PATH, read_only=True, data_only=True)
	worksheet = workbook[SHEET_NAME]
	header = None
	for values in worksheet.iter_rows(values_only=True):
		if header is None:
			header = values
			continue
		if any(values):
			yield dict(zip(header, values, strict=False))
	workbook.close()


def normalize_brand_key(brand):
	"""Collapse case/space/punctuation-only spelling variants for voting.

	Examples: LU/Lu, CandyLand/Candy Land, Mitchell's/Mitchells. This does not
	merge genuinely different names such as Pampers/Mother Comfort.
	"""
	return "".join(character for character in brand.casefold() if character.isalnum())


def collect_assignments(barcode_to_items, counts):
	"""Collect votes before writing so conflicting barcodes cannot silently make
	the final brand depend on spreadsheet row order. A strict majority wins;
	items with a tie are reported and skipped for manual review."""
	brand_votes = collections.defaultdict(collections.Counter)
	brand_spellings = collections.defaultdict(collections.Counter)
	spelling_last_seen = {}
	last_seen = {}
	unmatched_samples = []

	for index, row in enumerate(iter_rows(), start=1):
		counts["rows_seen"] += 1
		barcode = str(row.get("ItemCode") or "").strip()
		if len(barcode) < MIN_CODE_LEN:
			counts["skipped_short_code"] += 1
			continue

		brand = str(row.get("AiBrand") or "").strip()
		if not brand or brand.casefold() == "generic":
			counts["skipped_generic_brand"] += 1
			continue

		matched_items = barcode_to_items.get(barcode)
		if not matched_items:
			counts["skipped_unmatched_barcode"] += 1
			if len(unmatched_samples) < 50:
				unmatched_samples.append(barcode)
			continue

		counts["matched_workbook_rows"] += 1
		brand_key = normalize_brand_key(brand)
		brand_spellings[brand_key][brand] += len(matched_items)
		spelling_last_seen[(brand_key, brand)] = index
		if len(matched_items) > 1:
			counts["multi_match_barcodes"] += 1
		for item_name in matched_items:
			brand_votes[item_name][brand_key] += 1
			last_seen[(item_name, brand_key)] = index

	canonical_spelling = {}
	for brand_key, spellings in brand_spellings.items():
		ordered_spellings = sorted(
			spellings.items(),
			key=lambda pair: (pair[1], spelling_last_seen[(brand_key, pair[0])]),
			reverse=True,
		)
		canonical_spelling[brand_key] = ordered_spellings[0][0]
		if len(ordered_spellings) > 1:
			counts["normalized_spelling_variant_groups"] += 1

	assignments = {}
	conflicts = {}
	for item_name, votes in brand_votes.items():
		ordered = sorted(
			votes.items(),
			key=lambda pair: (pair[1], last_seen[(item_name, pair[0])]),
			reverse=True,
		)
		winner_key, winner_count = ordered[0]
		if len(ordered) > 1:
			conflicts[item_name] = [(canonical_spelling[brand_key], count) for brand_key, count in ordered]
			if winner_count == ordered[1][1]:
				counts["skipped_tied_brand_conflict"] += 1
				continue
			counts["resolved_brand_conflict_by_majority"] += 1
		assignments[item_name] = canonical_spelling[winner_key]

	counts["distinct_items_matched"] = len(brand_votes)
	counts["distinct_items_assignable"] = len(assignments)
	return assignments, conflicts, unmatched_samples


def canonicalize_brands(assignments, counts):
	existing = {str(name).casefold(): name for name in frappe.get_all("Brand", pluck="name")}
	canonical = {}
	for source_brand in sorted(set(assignments.values()), key=str.casefold):
		key = source_brand.casefold()
		brand_name = existing.get(key)
		if not brand_name:
			brand_name = source_brand
			counts["brands_would_create" if DRY_RUN else "brands_created"] += 1
			if not DRY_RUN:
				frappe.get_doc({"doctype": "Brand", "brand": brand_name}).insert(ignore_permissions=True)
			existing[key] = brand_name
		canonical[source_brand] = brand_name
	return canonical


def apply_assignments(assignments, canonical, counts):
	current = {
		row.name: (row.brand or "")
		for row in frappe.get_all(
			"Item", filters={"name": ["in", list(assignments)]}, fields=["name", "brand"]
		)
	}
	writes_since_commit = 0
	for item_name, source_brand in assignments.items():
		brand_name = canonical[source_brand]
		if current.get(item_name) == brand_name:
			counts["already_correct"] += 1
			continue
		if not DRY_RUN:
			frappe.db.set_value("Item", item_name, "brand", brand_name, update_modified=False)
		counts["items_would_update" if DRY_RUN else "items_updated"] += 1
		writes_since_commit += 1
		if not DRY_RUN and writes_since_commit >= COMMIT_EVERY:
			frappe.db.commit()
			writes_since_commit = 0
	if not DRY_RUN:
		frappe.db.commit()


def run():
	if frappe.local.site != TARGET_SITE:
		raise RuntimeError(f"This script is for {TARGET_SITE}, not {frappe.local.site}.")

	counts = collections.Counter()
	assignments, conflicts, unmatched_samples = collect_assignments(get_barcode_to_items(), counts)
	canonical = canonicalize_brands(assignments, counts)
	apply_assignments(assignments, canonical, counts)

	print("DRY_RUN", DRY_RUN)
	print("SUMMARY", dict(counts))
	print("DISTINCT_SOURCE_BRANDS", len(canonical))
	print("CONFLICTING_ITEMS", len(conflicts))
	print("CONFLICT_SAMPLE", list(conflicts.items())[:25])
	print("UNMATCHED_SAMPLE", unmatched_samples)


run()
